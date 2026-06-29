#!/usr/bin/env python3
"""
FOE Consolidated Duty Reporter
==============================
Scans an FOE Excel workbook for both G (Guard) duties and BDS duties for a given
person, across monthly sheets, and prints console statistics and saves data to
a JSON file for the React dashboard.

Usage:
    python foe_duty_reporter.py                        # prompts for file + name
    python foe_duty_reporter.py FOE_2026.xlsx          # prompts for name only
    python foe_duty_reporter.py FOE_2026.xlsx --name "ASHER LOW"  # fully automated

Options:
    --name NAME     Person's name
    --gen GEN       Optional GEN number to filter
    --start START   Start month (default: FEB25)
    --end END       End month (default: CURRENT)
    --json PATH     Path to save duties.json (default: duties.json)
    --out PATH      Path to save detailed CSV report
"""

from __future__ import annotations

import argparse
import calendar
import csv
import json
import os
import re
import sys
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook


# =========================
# HARD-SET DEFAULTS
# =========================
DEFAULT_WORKBOOK = "FOE 2026.xlsx"
DEFAULT_NAME = ""
DEFAULT_GEN = ""

# Range
DEFAULT_START = "FEB25"
DEFAULT_END = "CURRENT"   # Use "CURRENT" to scan until current month

# Output files
DEFAULT_JSON = "dashboard/public/duties.json"
DEFAULT_OUT = ""


MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}

MONTH_ABBR = {v: k for k, v in MONTHS.items()}


@dataclass(frozen=True)
class DutyHit:
    duty_date: date
    duty_raw: str
    duty_display: str
    duty_type: str  # "Guard" or "BDS"
    sheet: str
    row: int

    @property
    def day_name(self) -> str:
        return self.duty_date.strftime("%A")

    @property
    def category(self) -> str:
        weekday = self.duty_date.weekday()  # Monday = 0, Sunday = 6
        if weekday == 4:
            return "Friday"
        if weekday >= 5:
            return "Weekend"
        return "Weekday"

    @property
    def sort_key(self) -> tuple[int, date]:
        order = {
            "Weekday": 0,
            "Friday": 1,
            "Weekend": 2,
        }
        return order[self.category], self.duty_date


@dataclass(frozen=True)
class MonthResult:
    month: date
    sheet: str
    row: Optional[int]
    guard_count: int
    bds_count: int
    total: int
    matching_days: str
    status: str
    hits: tuple[DutyHit, ...] = field(default_factory=tuple)


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text.upper()


def normalize_name(value) -> str:
    return normalize_text(value)


def parse_month(value: str, *, allow_current: bool = False) -> date:
    raw = str(value or "").strip().upper()

    if allow_current and raw in {"CURRENT", "NOW", "TODAY"}:
        today = date.today()
        return date(today.year, today.month, 1)

    cleaned = raw.replace("_", " ").replace("-", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()

    match = re.fullmatch(r"([A-Z]{3,9})\s*(\d{2}|\d{4})", cleaned)
    if not match:
        raise ValueError(f"Invalid month '{value}'. Use format like JAN25 or JAN 2025.")

    month_word = match.group(1)[:3]
    if month_word not in MONTHS:
        raise ValueError(f"Invalid month name '{value}'.")

    month_num = MONTHS[month_word]
    year_num = int(match.group(2))
    year = 2000 + year_num if year_num < 100 else year_num

    return date(year, month_num, 1)


def month_label(month: date) -> str:
    return f"{MONTH_ABBR[month.month].title()} {month.year}"


def date_label(d: date) -> str:
    return f"{d.day} {MONTH_ABBR[d.month].title()} {d.year}"


def iter_months(start: date, end: date) -> Iterable[date]:
    y = start.year
    m = start.month

    while (y, m) <= (end.year, end.month):
        yield date(y, m, 1)
        m += 1
        if m == 13:
            y += 1
            m = 1


def build_month_sheet_map(workbook) -> dict[date, str]:
    result: dict[date, str] = {}
    for sheet_name in workbook.sheetnames:
        try:
            month = parse_month(sheet_name)
        except ValueError:
            continue
        result.setdefault(month, sheet_name)
    return result


def find_person_row(ws, target_name: str, target_gen: str = "") -> Optional[int]:
    wanted = normalize_name(target_name)
    target_gen = str(target_gen or "").strip()

    exact_matches: list[tuple[int, str]] = []
    loose_matches: list[tuple[int, str]] = []
    current_gen = ""

    for row_index, row in enumerate(
        ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=3,
            values_only=True,
        ),
        start=1,
    ):
        row_texts = [normalize_text(v) for v in row]
        for text in row_texts:
            gen_match = re.fullmatch(r"GEN\s*(\d+)", text)
            if gen_match:
                current_gen = gen_match.group(1)

        name_text = row_texts[2] if len(row_texts) >= 3 else ""
        if not name_text:
            continue

        if name_text == wanted:
            exact_matches.append((row_index, current_gen))
        elif wanted in name_text or name_text in wanted:
            loose_matches.append((row_index, current_gen))

    candidates = exact_matches or loose_matches
    if not candidates:
        return None

    if target_gen:
        for row_num, found_gen in candidates:
            if found_gen == target_gen:
                return row_num

    return candidates[0][0]


# Regex Patterns for G-duties and BDS-duties
GD_PATTERN = re.compile(r'G\s*\{d\}', re.IGNORECASE)
G_PLAIN_PATTERN = re.compile(r'(?<![A-Za-z])G(?!\s*[\{A-Za-z])')
BDS_PATTERN = re.compile(r'\bBDS\b', re.IGNORECASE)


def count_month(
    ws,
    month: date,
    row_num: int,
    sheet_name: str,
) -> tuple[int, int, str, tuple[DutyHit, ...]]:
    days_in_month = calendar.monthrange(month.year, month.month)[1]

    first_day_col = 4
    last_day_col = first_day_col + days_in_month - 1

    guard_count = 0
    bds_count = 0
    matching_days: list[str] = []
    hits: list[DutyHit] = []

    row_values = next(
        ws.iter_rows(
            min_row=row_num,
            max_row=row_num,
            min_col=first_day_col,
            max_col=last_day_col,
            values_only=True,
        ),
        [],
    )

    for day, value in enumerate(row_values, start=1):
        duty = normalize_text(value)
        duty_date = date(month.year, month.month, day)

        # 1. Guard Duty Check
        has_gd = bool(GD_PATTERN.search(duty))
        has_g = bool(G_PLAIN_PATTERN.search(duty))
        if has_gd or has_g:
            guard_count += 1
            display_val = "G {d}" if has_gd else "G"
            matching_days.append(f"{day}: {display_val}")
            hits.append(
                DutyHit(
                    duty_date=duty_date,
                    duty_raw=str(value).strip(),
                    duty_display=display_val,
                    duty_type="Guard",
                    sheet=sheet_name,
                    row=row_num,
                )
            )

        # 2. BDS Duty Check
        elif BDS_PATTERN.search(duty):
            bds_count += 1
            display_val = str(value).strip()
            matching_days.append(f"{day}: {display_val}")
            hits.append(
                DutyHit(
                    duty_date=duty_date,
                    duty_raw=display_val,
                    duty_display=display_val,
                    duty_type="BDS",
                    sheet=sheet_name,
                    row=row_num,
                )
            )

    return guard_count, bds_count, "; ".join(matching_days), tuple(hits)


def make_report(
    workbook_path: str,
    target_name: str,
    target_gen: str,
    start: date,
    end: date,
) -> list[MonthResult]:
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    month_to_sheet = build_month_sheet_map(wb)
    results: list[MonthResult] = []

    for month in iter_months(start, end):
        sheet_name = month_to_sheet.get(month, "")
        if not sheet_name:
            results.append(
                MonthResult(
                    month=month,
                    sheet="",
                    row=None,
                    guard_count=0,
                    bds_count=0,
                    total=0,
                    matching_days="",
                    status="Missing sheet",
                )
            )
            continue

        ws = wb[sheet_name]
        row_num = find_person_row(ws, target_name, target_gen)
        if row_num is None:
            results.append(
                MonthResult(
                    month=month,
                    sheet=sheet_name,
                    row=None,
                    guard_count=0,
                    bds_count=0,
                    total=0,
                    matching_days="",
                    status="Name not found",
                )
            )
            continue

        guard_count, bds_count, matching_days, hits = count_month(
            ws,
            month,
            row_num,
            sheet_name,
        )

        results.append(
            MonthResult(
                month=month,
                sheet=sheet_name,
                row=row_num,
                guard_count=guard_count,
                bds_count=bds_count,
                total=guard_count + bds_count,
                matching_days=matching_days,
                status="OK",
                hits=hits,
            )
        )

    wb.close()
    return results


def all_hits(results: list[MonthResult]) -> list[DutyHit]:
    hits: list[DutyHit] = []
    for item in results:
        hits.extend(item.hits)
    return sorted(hits, key=lambda h: h.sort_key)


def summary_counts(hits: list[DutyHit]) -> tuple[int, int, int]:
    weekdays = sum(1 for h in hits if h.category == "Weekday")
    fridays = sum(1 for h in hits if h.category == "Friday")
    weekends = sum(1 for h in hits if h.category == "Weekend")
    return weekdays, fridays, weekends


def print_console_report(
    results: list[MonthResult],
    target_name: str,
    target_gen: str,
    start: date,
    end: date,
) -> None:
    hits = all_hits(results)
    weekdays, fridays, weekends = summary_counts(hits)

    total_guard = sum(h.duty_type == "Guard" for h in hits)
    total_bds = sum(h.duty_type == "BDS" for h in hits)

    title_gen = f" / Gen {target_gen}" if target_gen else ""

    print("\n" + "=" * 75)
    print(f"  FOE Consolidated Duty Report")
    print(f"  Name : {target_name.upper()}{title_gen}")
    print(f"  Range: {month_label(start)} to {month_label(end)}")
    print("=" * 75)
    print(f"  Guard Duties: {total_guard:<15} BDS Duties: {total_bds:<15}")
    print(f"  Total Duties: {len(hits)}")
    print("-" * 75)
    print(f"  Weekdays (Mon-Thu): {weekdays:<10} Fridays: {fridays:<10} Weekends: {weekends:<10}")
    print("=" * 75 + "\n")

    print(f"  {'#':<4} {'Date':<16} {'Day':<12} {'Category':<14} {'Type':<8} {'Entry'}")
    print("  " + "-" * 71)

    if not hits:
        print("  No duties found.")
        return

    for idx, hit in enumerate(hits, start=1):
        print(
            f"  {idx:<4} {date_label(hit.duty_date):<16} {hit.day_name:<12} "
            f"{hit.category:<14} {hit.duty_type:<8} {hit.duty_display}"
        )
    print("\n" + "=" * 75 + "\n")


def write_csv(results: list[MonthResult], output_path: str) -> None:
    hits = all_hits(results)
    path = Path(output_path)
    if path.parent != Path("."):
        path.parent.mkdir(parents=True, exist_ok=True)

    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "number",
                "date",
                "day",
                "category",
                "duty_type",
                "duty_entry",
                "sheet",
                "row",
            ],
        )
        writer.writeheader()
        for idx, hit in enumerate(hits, start=1):
            writer.writerow(
                {
                    "number": idx,
                    "date": hit.duty_date.strftime("%Y-%m-%d"),
                    "day": hit.day_name,
                    "category": hit.category,
                    "duty_type": hit.duty_type,
                    "duty_entry": hit.duty_display,
                    "sheet": hit.sheet,
                    "row": hit.row,
                }
            )


def write_json_data(
    results: list[MonthResult],
    output_path: str,
    target_name: str,
    target_gen: str,
    start: date,
    end: date,
) -> None:
    hits = all_hits(results)
    weekdays, fridays, weekends = summary_counts(hits)

    total_guard = sum(h.duty_type == "Guard" for h in hits)
    total_bds = sum(h.duty_type == "BDS" for h in hits)

    payload = {
        "metadata": {
            "name": target_name.upper(),
            "gen": target_gen,
            "generated_at": datetime.now().isoformat(),
            "range": f"{month_label(start)} to {month_label(end)}",
            "start": start.strftime("%Y-%m-%d"),
            "end": end.strftime("%Y-%m-%d")
        },
        "summary": {
            "total_duties": len(hits),
            "total_guard": total_guard,
            "total_bds": total_bds,
            "weekdays": weekdays,
            "fridays": fridays,
            "weekends": weekends
        },
        "duties": [
            {
                "number": idx,
                "date": hit.duty_date.strftime("%Y-%m-%d"),
                "date_label": date_label(hit.duty_date),
                "month_label": month_label(hit.duty_date),
                "month_key": hit.duty_date.strftime("%b %y").upper(),
                "day": hit.day_name,
                "day_type": hit.category,
                "duty_type": hit.duty_type,
                "duty_display": hit.duty_display,
                "sheet": hit.sheet,
                "row": hit.row
            }
            for idx, hit in enumerate(hits, start=1)
        ]
    }

    path = Path(output_path)
    if path.parent != Path("."):
        path.parent.mkdir(parents=True, exist_ok=True)

    with path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Consolidate G and BDS duties from FOE monthly Excel sheets."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        default=DEFAULT_WORKBOOK,
        help="Path to FOE .xlsx file",
    )
    parser.add_argument(
        "--name",
        default=DEFAULT_NAME,
        help='Name to search for, e.g. "ASHER LOW"',
    )
    parser.add_argument(
        "--gen",
        default=DEFAULT_GEN,
        help='Optional GEN number, e.g. "21"',
    )
    parser.add_argument(
        "--start",
        default=DEFAULT_START,
        help="Start month, e.g. JAN25 or FEB 2025",
    )
    parser.add_argument(
        "--end",
        default=DEFAULT_END,
        help="End month, e.g. MAY26, or CURRENT",
    )
    parser.add_argument(
        "--out",
        default=DEFAULT_OUT,
        help="Optional detailed CSV output path",
    )
    parser.add_argument(
        "--json",
        default=DEFAULT_JSON,
        help="Optional JSON output path for React dashboard",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    workbook_path = args.workbook
    if not workbook_path:
        if not sys.stdin.isatty():
            workbook_path = DEFAULT_WORKBOOK
        else:
            workbook_path = input("Path to FOE .xlsx file: ").strip()

    target_name = args.name
    if not target_name:
        if not sys.stdin.isatty():
            raise SystemExit("Error: Name must be specified using --name in non-interactive environments.")
        else:
            target_name = input("Name to check, e.g. ASHER LOW: ").strip()

    if not workbook_path:
        raise SystemExit("No workbook path provided.")
    if not Path(workbook_path).exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    if not target_name:
        raise SystemExit("No name provided.")

    if args.name:
        target_gen = args.gen
    else:
        target_gen = args.gen or input("GEN number, optional. Press Enter to skip: ").strip()
    target_gen = str(target_gen).strip()

    start = parse_month(args.start)
    end = parse_month(args.end, allow_current=True)

    if start > end:
        raise SystemExit("Start month cannot be after end month.")

    print(f"\nScanning Excel workbook {Path(workbook_path).name} for {target_name.upper()}...")

    results = make_report(
        workbook_path=workbook_path,
        target_name=target_name,
        target_gen=target_gen,
        start=start,
        end=end,
    )

    print_console_report(
        results=results,
        target_name=target_name,
        target_gen=target_gen,
        start=start,
        end=end,
    )

    if args.out:
        write_csv(results, args.out)
        print(f"CSV report saved to: {Path(args.out).resolve()}")

    if args.json:
        write_json_data(
            results=results,
            output_path=args.json,
            target_name=target_name,
            target_gen=target_gen,
            start=start,
            end=end,
        )
        print(f"JSON data payload saved to: {Path(args.json).resolve()}")


if __name__ == "__main__":
    main()
