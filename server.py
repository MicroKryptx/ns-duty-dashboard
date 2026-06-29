#!/usr/bin/env python3
"""
FOE Duty Dashboard - optimized Flask API backend.

This version is tuned for small Railway/Render-style instances:
  - builds a compact duty index once instead of scanning Excel per request
  - serves stale cached data while Google Sheets refreshes in the background
  - exposes health/readiness endpoints for cloud probes
  - keeps the existing API routes compatible with the original dashboard
"""

from __future__ import annotations

import calendar
import json
import os
import re
import threading
import time
import urllib.request
from datetime import date, datetime
from pathlib import Path
from typing import Any

from flask import Flask, jsonify, request, send_from_directory
from flask_cors import CORS
from openpyxl import load_workbook

from foe_duty_reporter import (
    BDS_PATTERN,
    GD_PATTERN,
    G_PLAIN_PATTERN,
    MONTH_ABBR,
    date_label,
    iter_months,
    month_label,
    normalize_text,
    parse_month,
)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

BASE_DIR = Path(__file__).parent
DATA_DIR = Path(os.environ.get("DATA_DIR", BASE_DIR)).expanduser()
if not DATA_DIR.is_absolute():
    DATA_DIR = BASE_DIR / DATA_DIR
DATA_DIR.mkdir(parents=True, exist_ok=True)

GOOGLE_SHEET_ID = os.environ.get(
    "GOOGLE_SHEET_ID",
    "161I1PqF8qVl_ASUHiPUWYl5ortwthlJdggLufZUoG4o",
)
GOOGLE_EXPORT_URL = (
    f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/export?format=xlsx"
)

CACHE_FILE = Path(os.environ.get("WORKBOOK_CACHE_FILE", DATA_DIR / "_cached_foe_data.xlsx"))
INDEX_FILE = Path(os.environ.get("DUTY_INDEX_FILE", DATA_DIR / "_duty_index.json"))
LOCAL_FALLBACK = BASE_DIR / "FOE 2026.xlsx"
DASHBOARD_DIR = BASE_DIR / "dashboard" / "dist"

CACHE_TTL_SECONDS = int(os.environ.get("CACHE_TTL_SECONDS", "3600"))
DOWNLOAD_TIMEOUT_SECONDS = int(os.environ.get("DOWNLOAD_TIMEOUT_SECONDS", "12"))
STARTUP_BACKGROUND_REFRESH = (
    os.environ.get("STARTUP_BACKGROUND_REFRESH", "true").strip().lower()
    not in {"0", "false", "no", "off"}
)

DEFAULT_START = "FEB25"
INDEX_SCHEMA_VERSION = 1
SHEET_NAME_RE = re.compile(r"^[A-Z]{3}\d{2}$")


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def utcish_now_iso() -> str:
    return datetime.now().isoformat()


def month_key_for(month: date) -> str:
    return f"{MONTH_ABBR[month.month]}{month.year % 100:02d}"


def month_iso(month: date) -> str:
    return month.strftime("%Y-%m-%d")


def category_for_date(duty_date: date) -> str:
    weekday = duty_date.weekday()
    if weekday == 4:
        return "Friday"
    if weekday >= 5:
        return "Weekend"
    return "Weekday"


def category_sort_value(day_type: str) -> int:
    return {"Weekday": 0, "Friday": 1, "Weekend": 2}.get(day_type, 99)


def is_valid_number(value: Any) -> bool:
    if value is None:
        return False
    try:
        return int(value) > 0
    except (TypeError, ValueError):
        return False


def is_valid_rank(value: Any) -> bool:
    return value is not None and len(str(value).strip()) > 0


def is_valid_name(value: Any) -> bool:
    if value is None or not isinstance(value, str):
        return False
    name = value.strip().upper()
    return " " in name and len(name) >= 4


def file_fingerprint(path: Path) -> dict[str, Any]:
    stat = path.stat()
    return {
        "name": path.name,
        "size": stat.st_size,
        "mtime": stat.st_mtime,
    }


def read_json(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def atomic_write_json(path: Path, payload: dict[str, Any]) -> None:
    tmp_path = path.with_name(f"{path.name}.{os.getpid()}.{threading.get_ident()}.tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    tmp_path.replace(path)


def atomic_write_bytes(path: Path, payload: bytes) -> None:
    tmp_path = path.with_name(f"{path.name}.{os.getpid()}.{threading.get_ident()}.tmp")
    tmp_path.write_bytes(payload)
    tmp_path.replace(path)


# ---------------------------------------------------------------------------
# Duty index
# ---------------------------------------------------------------------------

class DutyNotReady(RuntimeError):
    pass


class DutyDataStore:
    def __init__(self) -> None:
        self._lock = threading.RLock()
        self._refresh_thread: threading.Thread | None = None
        self._index: dict[str, Any] | None = None
        self._refreshing = False
        self._last_error: str | None = None
        self._last_refresh_started_at: str | None = None
        self._last_refresh_finished_at: str | None = None

    # ----- state -----------------------------------------------------------

    def status(self) -> dict[str, Any]:
        with self._lock:
            index = self._index
            refreshing = self._refreshing
            error = self._last_error
            started = self._last_refresh_started_at
            finished = self._last_refresh_finished_at

        source_path = self._current_workbook_path()
        source = None
        last_refreshed = None
        age_seconds = None

        if index:
            source = index.get("source")
            last_refreshed = index.get("workbook_cached_at") or index.get("generated_at")
            age_seconds = self._age_from_iso(last_refreshed)
        elif source_path:
            source = "cached" if source_path == CACHE_FILE else "local_fallback"
            mtime = source_path.stat().st_mtime
            last_refreshed = datetime.fromtimestamp(mtime).isoformat()
            age_seconds = int(time.time() - mtime)

        return {
            "ready": index is not None,
            "refreshing": refreshing,
            "warming": index is None and refreshing,
            "last_refreshed": last_refreshed,
            "source": source,
            "age_seconds": age_seconds,
            "ttl_seconds": CACHE_TTL_SECONDS,
            "is_stale": age_seconds is not None and age_seconds > CACHE_TTL_SECONDS,
            "error": error,
            "last_refresh_started_at": started,
            "last_refresh_finished_at": finished,
        }

    def bootstrap(self) -> dict[str, Any]:
        self.ensure_background_refresh(force=False)
        with self._lock:
            index = self._index
            if not index:
                return {
                    "ready": False,
                    "names": [],
                    "months": [],
                    "data_status": self.status(),
                }
            return {
                "ready": True,
                "names": index["names"],
                "months": index["months"],
                "data_status": self.status(),
            }

    def names(self) -> list[str]:
        index = self._require_index()
        return list(index["names"])

    def months(self) -> list[dict[str, str]]:
        index = self._require_index()
        return list(index["months"])

    # ----- refresh ---------------------------------------------------------

    def load_persisted_index(self) -> bool:
        source_path = self._current_workbook_path()
        if not source_path or not INDEX_FILE.exists():
            return False
        loaded = self._load_index_if_current(source_path)
        if loaded:
            with self._lock:
                self._index = loaded
            return True
        return False

    def ensure_background_refresh(self, *, force: bool) -> bool:
        with self._lock:
            if self._refreshing:
                return False
            if not force and self._index is not None and not self.status()["is_stale"]:
                return False
            self._refreshing = True
            self._last_refresh_started_at = utcish_now_iso()
            self._refresh_thread = threading.Thread(
                target=self._refresh_worker,
                kwargs={"force": force},
                name="duty-data-refresh",
                daemon=True,
            )
            self._refresh_thread.start()
            return True

    def refresh_blocking(self, *, force: bool) -> dict[str, Any]:
        with self._lock:
            if self._refreshing:
                thread = self._refresh_thread
            else:
                self._refreshing = True
                self._last_refresh_started_at = utcish_now_iso()
                thread = None

        if thread:
            thread.join(timeout=DOWNLOAD_TIMEOUT_SECONDS + 30)
        else:
            self._refresh_worker(force=force)

        return self.status()

    def _refresh_worker(self, *, force: bool) -> None:
        try:
            self._load_or_build_from_available_source(source_hint=None)

            should_download = force or self.status()["is_stale"] or not CACHE_FILE.exists()
            if should_download:
                result = self._download_workbook(force=force)
                if result["status"] == "ok":
                    self._load_or_build_from_available_source(source_hint="google_sheets")
                elif self._index is None:
                    self._load_or_build_from_available_source(source_hint=None)
        except Exception as exc:
            with self._lock:
                self._last_error = str(exc)
        finally:
            with self._lock:
                self._refreshing = False
                self._last_refresh_finished_at = utcish_now_iso()

    def _download_workbook(self, *, force: bool) -> dict[str, Any]:
        if not force and CACHE_FILE.exists() and not self.status()["is_stale"]:
            return {"status": "cached", "message": "Workbook cache is fresh."}

        try:
            req = urllib.request.Request(
                GOOGLE_EXPORT_URL,
                headers={"User-Agent": "FOE-Duty-Dashboard/2.0"},
            )
            with urllib.request.urlopen(req, timeout=DOWNLOAD_TIMEOUT_SECONDS) as response:
                payload = response.read()

            if payload[:2] != b"PK":
                raise ValueError("Downloaded response is not a valid .xlsx file.")

            atomic_write_bytes(CACHE_FILE, payload)
            with self._lock:
                self._last_error = None
            return {
                "status": "ok",
                "message": f"Downloaded workbook from Google Sheets ({len(payload) // 1024} KB).",
            }
        except Exception as exc:
            with self._lock:
                self._last_error = str(exc)
            return {
                "status": "fallback_cached" if CACHE_FILE.exists() else "error",
                "message": f"Google Sheets refresh failed: {exc}",
                "error": str(exc),
            }

    # ----- index build/load ------------------------------------------------

    def _load_or_build_from_available_source(self, source_hint: str | None) -> None:
        source_path = self._current_workbook_path()
        if not source_path:
            raise FileNotFoundError(
                "No workbook data is available. Add _cached_foe_data.xlsx, FOE 2026.xlsx, "
                "or allow Google Sheets refresh to succeed."
            )

        loaded = self._load_index_if_current(source_path)
        if loaded:
            with self._lock:
                self._index = loaded
            return

        source = source_hint or ("cached" if source_path == CACHE_FILE else "local_fallback")
        index = self._build_index(source_path, source)
        atomic_write_json(INDEX_FILE, index)
        with self._lock:
            self._index = index

    def _load_index_if_current(self, source_path: Path) -> dict[str, Any] | None:
        if not INDEX_FILE.exists():
            return None
        try:
            index = read_json(INDEX_FILE)
            if index.get("schema_version") != INDEX_SCHEMA_VERSION:
                return None
            if index.get("workbook_fingerprint") != file_fingerprint(source_path):
                return None
            return index
        except Exception:
            return None

    def _build_index(self, workbook_path: Path, source: str) -> dict[str, Any]:
        started = time.perf_counter()
        workbook_fingerprint = file_fingerprint(workbook_path)
        workbook_cached_at = datetime.fromtimestamp(workbook_fingerprint["mtime"]).isoformat()

        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            month_to_sheet: dict[date, str] = {}
            for sheet_name in wb.sheetnames:
                if not SHEET_NAME_RE.match(sheet_name):
                    continue
                try:
                    month_to_sheet.setdefault(parse_month(sheet_name), sheet_name)
                except ValueError:
                    continue

            months = [
                {
                    "key": month_key_for(month),
                    "label": month_label(month),
                    "date": month_iso(month),
                }
                for month in sorted(month_to_sheet)
            ]

            people: dict[str, dict[str, Any]] = {}
            total_hits = 0

            for month in sorted(month_to_sheet):
                sheet_name = month_to_sheet[month]
                ws = wb[sheet_name]
                current_gen = ""
                days_in_month = calendar.monthrange(month.year, month.month)[1]
                month_id = month_iso(month)

                day_meta = []
                for day in range(1, days_in_month + 1):
                    duty_date = date(month.year, month.month, day)
                    day_type = category_for_date(duty_date)
                    day_meta.append(
                        {
                            "day": day,
                            "date": duty_date.strftime("%Y-%m-%d"),
                            "date_label": date_label(duty_date),
                            "month_label": month_label(duty_date),
                            "month_key": duty_date.strftime("%b %y").upper(),
                            "day_name": duty_date.strftime("%A"),
                            "day_type": day_type,
                            "sort_day_type": category_sort_value(day_type),
                        }
                    )

                for row_num, row_values in enumerate(
                    ws.iter_rows(min_row=1, max_col=34, values_only=True),
                    start=1,
                ):
                    first = row_values[0] if len(row_values) > 0 else None
                    rank = row_values[1] if len(row_values) > 1 else None
                    raw_name = row_values[2] if len(row_values) > 2 else None

                    for value in (first, rank, raw_name):
                        text = normalize_text(value)
                        if text.startswith("GEN"):
                            gen_match = re.fullmatch(r"GEN\s*(\d+)", text)
                            if gen_match:
                                current_gen = gen_match.group(1)

                    if not (
                        is_valid_number(first)
                        and is_valid_rank(rank)
                        and is_valid_name(raw_name)
                    ):
                        continue

                    person_name = normalize_text(raw_name)
                    person = people.setdefault(
                        person_name,
                        {
                            "display_name": person_name,
                            "months": {},
                        },
                    )
                    entries = person["months"].setdefault(month_id, [])

                    hits: list[dict[str, Any]] = []
                    day_cells = row_values[3 : 3 + days_in_month]
                    for meta, value in zip(day_meta, day_cells):
                        duty = normalize_text(value)
                        if not duty:
                            continue

                        duty_type = ""
                        display_value = ""
                        has_gd = bool(GD_PATTERN.search(duty))
                        has_g = bool(G_PLAIN_PATTERN.search(duty))
                        if has_gd or has_g:
                            duty_type = "Guard"
                            display_value = "G {d}" if has_gd else "G"
                        elif BDS_PATTERN.search(duty):
                            duty_type = "BDS"
                            display_value = str(value).strip()

                        if not duty_type:
                            continue

                        hits.append(
                            {
                                "date": meta["date"],
                                "date_label": meta["date_label"],
                                "month_label": meta["month_label"],
                                "month_key": meta["month_key"],
                                "day": meta["day_name"],
                                "day_type": meta["day_type"],
                                "duty_type": duty_type,
                                "duty_display": display_value,
                                "sheet": sheet_name,
                                "row": row_num,
                                "_sort_day_type": meta["sort_day_type"],
                            }
                        )

                    total_hits += len(hits)
                    entries.append(
                        {
                            "gen": current_gen,
                            "sheet": sheet_name,
                            "row": row_num,
                            "hits": hits,
                        }
                    )

            elapsed = time.perf_counter() - started
            names = sorted(people.keys())
            return {
                "schema_version": INDEX_SCHEMA_VERSION,
                "generated_at": utcish_now_iso(),
                "source": source,
                "workbook_cached_at": workbook_cached_at,
                "workbook_fingerprint": workbook_fingerprint,
                "stats": {
                    "build_seconds": round(elapsed, 3),
                    "people": len(names),
                    "months": len(months),
                    "hits": total_hits,
                },
                "names": names,
                "months": months,
                "people": people,
            }
        finally:
            wb.close()

    # ----- query -----------------------------------------------------------

    def validate_person(
        self,
        target_name: str,
        start: date,
        end: date,
        target_gen: str = "",
    ) -> dict[str, Any]:
        index = self._require_index()
        resolved_name = self._resolve_name(index, target_name)
        found_in: list[str] = []
        missing_in: list[str] = []

        if not resolved_name:
            for month in iter_months(start, end):
                missing_in.append(month_label(month))
            return {
                "found": False,
                "found_in": found_in,
                "missing_in": missing_in,
                "total_sheets_checked": len(missing_in),
            }

        person = index["people"][resolved_name]
        for month in iter_months(start, end):
            entry = self._select_entry(person, month, target_gen)
            if entry:
                found_in.append(month_label(month))
            else:
                missing_in.append(month_label(month))

        return {
            "found": bool(found_in),
            "found_in": found_in,
            "missing_in": missing_in,
            "total_sheets_checked": len(found_in) + len(missing_in),
        }

    def build_duty_payload(
        self,
        target_name: str,
        target_gen: str,
        start: date,
        end: date,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        index = self._require_index()
        resolved_name = self._resolve_name(index, target_name)
        validation = self.validate_person(target_name, start, end, target_gen)

        if not resolved_name or not validation["found"]:
            return (
                {
                    "valid": False,
                    "error": (
                        f"{target_name} was not found in any sheet between "
                        f"{month_label(start)} and {month_label(end)}."
                    ),
                    "details": validation,
                    "data_status": self.status(),
                },
                validation,
            )

        person = index["people"][resolved_name]
        hits: list[dict[str, Any]] = []
        for month in iter_months(start, end):
            entry = self._select_entry(person, month, target_gen)
            if entry:
                hits.extend(entry["hits"])

        hits = sorted(hits, key=lambda hit: (hit["_sort_day_type"], hit["date"]))
        duties = []
        for number, hit in enumerate(hits, start=1):
            public_hit = {key: value for key, value in hit.items() if not key.startswith("_")}
            public_hit["number"] = number
            duties.append(public_hit)

        total_guard = sum(hit["duty_type"] == "Guard" for hit in hits)
        total_bds = sum(hit["duty_type"] == "BDS" for hit in hits)
        weekdays = sum(hit["day_type"] == "Weekday" for hit in hits)
        fridays = sum(hit["day_type"] == "Friday" for hit in hits)
        weekends = sum(hit["day_type"] == "Weekend" for hit in hits)

        payload = {
            "valid": True,
            "metadata": {
                "name": resolved_name,
                "gen": target_gen,
                "generated_at": utcish_now_iso(),
                "range": f"{month_label(start)} to {month_label(end)}",
                "start": start.strftime("%Y-%m-%d"),
                "end": end.strftime("%Y-%m-%d"),
            },
            "summary": {
                "total_duties": len(hits),
                "total_guard": total_guard,
                "total_bds": total_bds,
                "weekdays": weekdays,
                "fridays": fridays,
                "weekends": weekends,
            },
            "duties": duties,
            "details": validation,
            "data_status": self.status(),
        }
        return payload, validation

    def _require_index(self) -> dict[str, Any]:
        with self._lock:
            if not self._index:
                raise DutyNotReady("Duty data is still warming up. Please retry shortly.")
            return self._index

    def _resolve_name(self, index: dict[str, Any], target_name: str) -> str | None:
        wanted = normalize_text(target_name)
        if wanted in index["people"]:
            return wanted

        candidates = [
            name for name in index["names"]
            if wanted and (wanted in name or name in wanted)
        ]
        return candidates[0] if candidates else None

    def _select_entry(
        self,
        person: dict[str, Any],
        month: date,
        target_gen: str,
    ) -> dict[str, Any] | None:
        entries = person["months"].get(month_iso(month), [])
        if not entries:
            return None

        target_gen = str(target_gen or "").strip()
        if target_gen:
            for entry in entries:
                if str(entry.get("gen", "")).strip() == target_gen:
                    return entry

        return entries[0]

    def _current_workbook_path(self) -> Path | None:
        if CACHE_FILE.exists():
            return CACHE_FILE
        if LOCAL_FALLBACK.exists():
            return LOCAL_FALLBACK
        return None

    @staticmethod
    def _age_from_iso(value: str | None) -> int | None:
        if not value:
            return None
        try:
            return int(time.time() - datetime.fromisoformat(value).timestamp())
        except ValueError:
            return None


store = DutyDataStore()
store.load_persisted_index()


# ---------------------------------------------------------------------------
# Flask app
# ---------------------------------------------------------------------------

app = Flask(__name__, static_folder=str(DASHBOARD_DIR), static_url_path="")
CORS(app)


def parse_range_from_request(body: dict[str, Any]) -> tuple[date, date]:
    start = parse_month(body.get("start", DEFAULT_START))
    end = parse_month(body.get("end", "CURRENT"), allow_current=True)
    if start > end:
        raise ValueError("Start month cannot be after end month.")
    return start, end


def not_ready_response():
    return jsonify({
        "ready": False,
        "error": "Duty data is still warming up. Please retry shortly.",
        "data_status": store.status(),
    }), 503


@app.route("/api/bootstrap", methods=["GET"])
def bootstrap():
    return jsonify(store.bootstrap())


@app.route("/api/names", methods=["GET"])
def get_names():
    try:
        return jsonify({"names": store.names(), "data_status": store.status()})
    except DutyNotReady:
        return not_ready_response()
    except Exception as exc:
        return jsonify({"error": str(exc), "data_status": store.status()}), 500


@app.route("/api/months", methods=["GET"])
def get_months():
    try:
        return jsonify({"months": store.months(), "data_status": store.status()})
    except DutyNotReady:
        return not_ready_response()
    except Exception as exc:
        return jsonify({"error": str(exc), "data_status": store.status()}), 500


@app.route("/api/validate", methods=["POST"])
def validate_person():
    body = request.get_json(silent=True) or {}
    name = body.get("name", "").strip()
    if not name:
        return jsonify({"error": "No name provided."}), 400

    try:
        start, end = parse_range_from_request(body)
        result = store.validate_person(name, start, end, body.get("gen", ""))
        if not result["found"]:
            return jsonify({
                "valid": False,
                "error": (
                    f"{name} was not found in any sheet between "
                    f"{month_label(start)} and {month_label(end)}."
                ),
                "details": result,
                "data_status": store.status(),
            })
        return jsonify({"valid": True, "details": result, "data_status": store.status()})
    except DutyNotReady:
        return not_ready_response()
    except ValueError as exc:
        return jsonify({"error": str(exc), "data_status": store.status()}), 400
    except Exception as exc:
        return jsonify({"error": str(exc), "data_status": store.status()}), 500


@app.route("/api/duties", methods=["POST"])
def get_duties():
    body = request.get_json(silent=True) or {}
    name = body.get("name", "").strip()
    if not name:
        return jsonify({"error": "No name provided."}), 400

    try:
        start, end = parse_range_from_request(body)
        payload, _validation = store.build_duty_payload(
            target_name=name,
            target_gen=body.get("gen", ""),
            start=start,
            end=end,
        )
        return jsonify(payload)
    except DutyNotReady:
        return not_ready_response()
    except ValueError as exc:
        return jsonify({"error": str(exc), "data_status": store.status()}), 400
    except Exception as exc:
        return jsonify({"error": str(exc), "data_status": store.status()}), 500


@app.route("/api/refresh", methods=["POST"])
def refresh_data():
    wait = request.args.get("wait", "").strip().lower() in {"1", "true", "yes"}
    if wait:
        status = store.refresh_blocking(force=True)
        return jsonify({
            "status": "ok" if status["ready"] else "error",
            "message": "Refresh completed." if status["ready"] else status.get("error"),
            "data_status": status,
        })

    started = store.ensure_background_refresh(force=True)
    return jsonify({
        "status": "refreshing" if started else "already_refreshing",
        "message": (
            "Refresh started in the background."
            if started
            else "Refresh is already running in the background."
        ),
        "data_status": store.status(),
    })


@app.route("/api/status", methods=["GET"])
def data_status():
    store.ensure_background_refresh(force=False)
    return jsonify(store.status())


@app.route("/api/healthz", methods=["GET"])
def healthz():
    return jsonify({"ok": True})


@app.route("/api/readyz", methods=["GET"])
def readyz():
    status = store.status()
    code = 200 if status["ready"] else 503
    return jsonify(status), code


@app.route("/", defaults={"path": ""})
@app.route("/<path:path>")
def serve_react(path: str):
    file_path = DASHBOARD_DIR / path
    if path and file_path.exists():
        response = send_from_directory(str(DASHBOARD_DIR), path)
        if path.startswith("assets/"):
            response.headers["Cache-Control"] = "public, max-age=31536000, immutable"
        else:
            response.headers["Cache-Control"] = "public, max-age=3600"
        return response

    response = send_from_directory(str(DASHBOARD_DIR), "index.html")
    response.headers["Cache-Control"] = "no-cache"
    return response


if STARTUP_BACKGROUND_REFRESH:
    store.ensure_background_refresh(force=False)


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    debug = os.environ.get("FLASK_DEBUG", "false").strip().lower() == "true"
    print(f"\n  Google Sheet: {GOOGLE_SHEET_ID}")
    print(f"  Data Dir:     {DATA_DIR}")
    print(f"  Cache File:   {CACHE_FILE}")
    print(f"  Index File:   {INDEX_FILE}")
    print(f"  Dashboard:    {DASHBOARD_DIR}")
    print(f"  Starting at:  http://localhost:{port}\n")
    app.run(host="0.0.0.0", port=port, debug=debug)
